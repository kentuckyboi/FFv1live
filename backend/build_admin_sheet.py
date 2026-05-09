"""Build the First Five admin Sheet template (.xlsx).
   Tabs: README | Members | Completions | Live Progress | Tasks Reference
   Live Progress is the formula-driven dashboard staff watch during conference.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

OUT = "/sessions/admiring-practical-tesla/mnt/outputs/first-five-app/backend/admin-sheet-template.xlsx"

BLUE = "1B365D"
BLUE_DEEP = "0F2342"
GOLD = "C9A961"
ORANGE = "E87722"
CREAM = "FAF7F2"
LINE = "E6E3DC"
GOOD = "2E7D5B"
WHITE = "FFFFFF"

wb = Workbook()

def header_fill(cell, color=BLUE, font_color=WHITE):
    cell.fill = PatternFill("solid", start_color=color)
    cell.font = Font(name="Arial", size=11, bold=True, color=font_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---------- Tab 1: README ----------
ws = wb.active
ws.title = "README"
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 110

ws['B2'] = "First Five Project — Admin Sheet"
ws['B2'].font = Font(name="Arial", size=22, bold=True, color=BLUE)
ws['B3'] = "Live progress tracker for the National Conference"
ws['B3'].font = Font(name="Arial", size=12, italic=True, color=ORANGE)

ws['B5'] = "How this sheet is wired up"
ws['B5'].font = Font(name="Arial", size=14, bold=True, color=BLUE)

readme_lines = [
    "1.  This is the SHEET tab structure expected by the Apps Script backend (Code.gs).",
    "2.  Upload this file to Google Drive, open it as a Google Sheet, then follow backend/SETUP.md to attach Code.gs.",
    "3.  The Apps Script will write to 'Members' (when someone signs up) and 'Completions' (when they finish a task).",
    "4.  'Live Progress' is formula-driven — it auto-summarizes Members + Completions. You don't write to it; just watch it during the conference.",
    "5.  Refresh frequency: Google Sheets recalculates automatically every few seconds when data arrives. Your view will update live.",
    "6.  During conference: leave Live Progress open on a second monitor. Sort by 'Last Activity' to see who just earned a badge.",
    "7.  Privacy: Only people you explicitly share the Sheet with will see member data. Don't share the link publicly.",
    "8.  Photos: signature and photo blobs are stored in 'Completions' as base64 strings (only photos under ~200KB). Click a cell to peek.",
]
for i, line in enumerate(readme_lines, start=6):
    c = ws.cell(row=i, column=2, value=line)
    c.font = Font(name="Arial", size=11, color="303644")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 26

ws['B15'] = "Pro tip"
ws['B15'].font = Font(name="Arial", size=14, bold=True, color=ORANGE)
ws['B16'] = ("Build a tiny pivot or chart on Live Progress (Insert > Chart in Google Sheets) showing percent completion by chapter. "
             "Project it on a screen at the back of the registration table. Members love seeing their chapter climb.")
ws['B16'].font = Font(name="Arial", size=11, italic=True, color="303644")
ws['B16'].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[16].height = 60

# ---------- Tab 2: Members ----------
ws = wb.create_sheet("Members")
m_headers = ["email", "name", "chapter", "year", "headline", "source", "created_at"]
widths = [30, 22, 22, 8, 50, 14, 22]
for i, (h, w) in enumerate(zip(m_headers, widths), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    cell = ws.cell(row=1, column=i, value=h)
    header_fill(cell)
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

# Seed with the demo roster
seed = [
    ("tyrone@example.org", "Tyrone Gentry", "Texas Alpha", 2,
     "I'd love to talk to someone about: 4-H program design.", "seed", ""),
    ("marisol@example.org", "Marisol Vega", "California Beta", 1,
     "I'd love to talk to someone about: rural broadband adoption.", "seed", ""),
    ("denzel@example.org", "Denzel Carter", "Georgia Gamma", 3,
     "I'd love to talk to someone about: livestock economics.", "seed", ""),
    ("fiona@example.org", "Fiona O'Connell", "Iowa Delta", 2,
     "I'd love to talk to someone about: nutrition curriculum for K-5.", "seed", ""),
    ("jamal@example.org", "Jamal Whitfield", "Florida Epsilon", 4,
     "I'd love to talk to someone about: state-fair extension exhibits.", "seed", ""),
    ("anya@example.org", "Anya Okonkwo", "New York Zeta", 1,
     "I'd love to talk to someone about: climate-smart small farms.", "seed", ""),
]
for r, row in enumerate(seed, start=2):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", size=10, color="000000")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if r % 2 == 0:
        for c in range(1, len(m_headers) + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", start_color=CREAM)
    ws.row_dimensions[r].height = 32

# ---------- Tab 3: Completions ----------
ws = wb.create_sheet("Completions")
c_headers = ["email", "task_id", "proof_kind", "validator", "answer", "photo_data_url", "completed_at"]
widths = [30, 12, 14, 22, 30, 16, 22]
for i, (h, w) in enumerate(zip(c_headers, widths), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    cell = ws.cell(row=1, column=i, value=h)
    header_fill(cell)
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

# (No seed — it fills as people complete tasks)
ws.cell(row=2, column=1, value="(Empty until a member completes their first task)")
ws.cell(row=2, column=1).font = Font(name="Arial", size=10, italic=True, color="909090")

# ---------- Tab 4: Live Progress (the dashboard) ----------
ws = wb.create_sheet("Live Progress")

ws['A1'] = "Live Progress Dashboard"
ws['A1'].font = Font(name="Arial", size=18, bold=True, color=BLUE)
ws.merge_cells('A1:G1')

ws['A2'] = "This sheet recalculates automatically as Members and Completions are written by the app."
ws['A2'].font = Font(name="Arial", size=10, italic=True, color="606060")
ws.merge_cells('A2:G2')

# Headers (row 4)
lp_headers = ["Email", "Name", "Chapter", "Year", "Tasks Done", "Points", "Tier", "Last Activity"]
widths = [30, 22, 20, 7, 12, 9, 10, 22]
for i, (h, w) in enumerate(zip(lp_headers, widths), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    cell = ws.cell(row=4, column=i, value=h)
    header_fill(cell)
ws.row_dimensions[4].height = 28
ws.freeze_panes = "A5"

# Formulas — built to scale to 500 members and 5000 completions.
MEMBERS_RANGE = "Members!A2:G500"
COMPLETIONS_RANGE = "Completions!A2:G5000"

# Per-task points (used by SUMPRODUCT to compute points per email)
# Build a points lookup using a hidden helper region in this sheet.
# Columns J..K hold task_id and points; we'll write that block first.
# Tasks reference (hidden helper)
ws.column_dimensions['J'].hidden = True
ws.column_dimensions['K'].hidden = True
ws['J3'] = "task_id"; ws['K3'] = "points"
ws['J3'].font = Font(bold=True); ws['K3'].font = Font(bold=True)
TASK_POINTS = [
    ("T01", 3), ("T02", 3), ("T03", 2), ("T04", 2), ("T05", 2),
    ("T06", 4), ("T07", 3), ("T08", 3), ("T09", 3), ("T10", 2),
    ("T11", 2), ("T12", 4), ("T13", 2), ("T14", 3), ("T15", 2),
]
for i, (tid, pts) in enumerate(TASK_POINTS, start=4):
    ws.cell(row=i, column=10, value=tid)
    ws.cell(row=i, column=11, value=pts)

# 500 progress rows mirroring Members
for i in range(5, 505):
    member_row = i - 3  # Members!A2 maps to dashboard row 5
    # A: email
    ws.cell(row=i, column=1, value=f'=IFERROR(IF(Members!A{member_row}="","",Members!A{member_row}),"")')
    ws.cell(row=i, column=2, value=f'=IFERROR(IF(Members!A{member_row}="","",Members!B{member_row}),"")')
    ws.cell(row=i, column=3, value=f'=IFERROR(IF(Members!A{member_row}="","",Members!C{member_row}),"")')
    ws.cell(row=i, column=4, value=f'=IFERROR(IF(Members!A{member_row}="","",Members!D{member_row}),"")')
    # E: tasks done
    ws.cell(row=i, column=5,
            value=f'=IF(A{i}="","",COUNTIF(Completions!A:A,A{i}))')
    # F: total points = sum over completions where email=A{i} of vlookup(task_id) into helper J:K
    ws.cell(row=i, column=6,
            value=(f'=IF(A{i}="","",'
                   f'SUMPRODUCT((Completions!A$2:A$5000=A{i})*'
                   f'IFERROR(VLOOKUP(Completions!B$2:B$5000,$J$4:$K$18,2,FALSE),0)))'))
    # G: tier
    ws.cell(row=i, column=7,
            value=(f'=IF(A{i}="","",IF(F{i}>=20,"Gold",IF(F{i}>=14,"Silver",IF(F{i}>=8,"Bronze","—"))))'))
    # H: last activity
    ws.cell(row=i, column=8,
            value=(f'=IF(A{i}="","",IFERROR(TEXT(MAXIFS(Completions!G:G,Completions!A:A,A{i}),"yyyy-mm-dd hh:mm"),"—"))'))

    # Row formatting
    for c in range(1, 9):
        cell = ws.cell(row=i, column=c)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
    if i % 2 == 0:
        for c in range(1, 9):
            ws.cell(row=i, column=c).fill = PatternFill("solid", start_color=CREAM)
    ws.row_dimensions[i].height = 22

# Conditional formatting: tier coloring (column G)
gold_fill = PatternFill("solid", start_color="FFE9B0")
silver_fill = PatternFill("solid", start_color="E8E8E8")
bronze_fill = PatternFill("solid", start_color="F4DDC1")
ws.conditional_formatting.add("G5:G504",
    CellIsRule(operator="equal", formula=['"Gold"'], fill=gold_fill))
ws.conditional_formatting.add("G5:G504",
    CellIsRule(operator="equal", formula=['"Silver"'], fill=silver_fill))
ws.conditional_formatting.add("G5:G504",
    CellIsRule(operator="equal", formula=['"Bronze"'], fill=bronze_fill))

# Conditional formatting: heat-map on Points column (F)
ws.conditional_formatting.add("F5:F504",
    ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                   mid_type='num', mid_value=10, mid_color='FFE5C2',
                   end_type='num', end_value=20, end_color=ORANGE))

# Quick stats block at the top
ws['I1'] = "Conference Snapshot"
ws['I1'].font = Font(name="Arial", size=14, bold=True, color=BLUE)
ws.column_dimensions['I'].width = 22
ws.column_dimensions['L'].width = 14

stats = [
    ("Total members",          '=COUNTIF(A5:A504,"?*")'),
    ("Members with ≥1 task",   '=COUNTIF(E5:E504,">0")'),
    ("Tasks completed total",  '=SUM(E5:E504)'),
    ("Bronze tier",            '=COUNTIF(G5:G504,"Bronze")'),
    ("Silver tier",            '=COUNTIF(G5:G504,"Silver")'),
    ("Gold tier",              '=COUNTIF(G5:G504,"Gold")'),
]
for i, (label, formula) in enumerate(stats, start=2):
    ws.cell(row=i, column=9, value=label).font = Font(name="Arial", size=10, bold=True, color=BLUE)
    ws.cell(row=i, column=12, value=formula).font = Font(name="Arial", size=11, color=ORANGE)
    ws.cell(row=i, column=12).alignment = Alignment(horizontal="right")

# ---------- Tab 5: Tasks Reference (read-only mirror) ----------
ws = wb.create_sheet("Tasks Reference")
t_headers = ["task_id", "task_name", "category", "proof_type", "points", "where", "mission_tie_in"]
widths = [10, 38, 22, 14, 8, 28, 38]
for i, (h, w) in enumerate(zip(t_headers, widths), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    cell = ws.cell(row=1, column=i, value=h)
    header_fill(cell)

tasks_seed = [
    ("T01", "Attend the First-Timer's Welcome", "Conference Anchor", "Signature", 3, "Main Hall · Day 1, 9:00 AM", "Article XI – Creed: shared experience and fellowship."),
    ("T02", "Meet a National Board Member", "Networking", "Signature", 3, "Networking Lounge · Days 1-3", "Strategic Goal 1: Increase value of ESP National to chapters."),
    ("T03", "Visit the Poster Session", "Professional Development", "Photo", 2, "Hall B · Day 2", "Strategic Goal 2: Excellence through professional development."),
    ("T04", "Selfie at the Centennial Countdown Wall", "ESP Identity", "Photo", 2, "Lobby · Anytime", "ESP heritage and the path to 100 years."),
    ("T05", "Knowledge Check: Opening Keynote", "Knowledge", "Quiz", 2, "After Day 1 keynote", "Strategic Goal 2: Professional development."),
    ("T06", "Introduce Yourself to Three New Chapters", "Networking", "Signature", 4, "Anywhere · By end of Day 2", "Strategic Goal 1: Connect chapters across the country."),
    ("T07", "Attend a Committee Meeting", "Professional Development", "Signature", 3, "See conference program · Day 2", "Article I – Mission: governance through committees."),
    ("T08", "Knowledge Check: A Scholarship I Could Apply For", "Knowledge", "Quiz", 3, "After Awards & Scholarships booth", "Strategic Goal 2: Development opportunities."),
    ("T09", "Photo with a Mentor", "Networking", "Photo", 3, "Anywhere · Days 1-3", "Article XI – Creed: 'share with others the lessons of my experience.'"),
    ("T10", "Visit Your Regional Caucus", "ESP Identity", "Signature", 2, "See conference program · Day 2", "Strategic Goal: Enhance chapter leader engagement."),
    ("T11", "Knowledge Check: First-Timer's Session", "Knowledge", "Quiz", 2, "After the First-Timer's session, Day 1", "Article XI – Creed: fellowship and shared learning."),
    ("T12", "Sign Up for a Committee for Next Year", "Engagement", "Signature", 4, "Membership table · Days 2-3", "Strategic Goal 1: Connect national committees to chapter units."),
    ("T13", "Attend a Session Outside Your Discipline", "Professional Development", "Photo", 2, "Any breakout · Days 1-3", "Strategic Goal 2: Multi-disciplinary professional development."),
    ("T14", "Closing Banquet Attendance", "Conference Anchor", "Signature", 3, "Banquet Hall · Final evening", "ESP heritage and recognition tradition."),
    ("T15", "Reflection: What Will You Bring Home?", "Engagement", "Quiz", 2, "Last day, after closing remarks", "Article XI – Creed: 'apply lessons learned' to my home work."),
]
for r, row in enumerate(tasks_seed, start=2):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    if r % 2 == 0:
        for c in range(1, len(t_headers) + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", start_color=CREAM)
    ws.row_dimensions[r].height = 32

wb.save(OUT)
print(f"Saved: {OUT}")
